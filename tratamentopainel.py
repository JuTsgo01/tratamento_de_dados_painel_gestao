# -*- coding: utf-8 -*-
"""tratamentopainel.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1PsFS4iQQfF4GyAvzJa4ZWKNK_X2jkXqq

Importando as bibliotecas necessárias
"""

import pandas as pd
import openpyxl
from datetime import datetime, timedelta

"""Definindo a data inicial e final do painel

"""

data_inicial = datetime.now() - timedelta(days=8)
data_inicial_formatada = data_inicial.strftime('%Y-%m-%d')

data_final = datetime.now() - timedelta(days=2)
data_final_formatada = data_final.strftime('%Y-%m-%d')

data_para_nome_planilha = datetime.now().strftime('%Y-%m-%d')

nome_arquivo = (f'/content/drive/MyDrive/painel de gestão/Filtro para subir no phpmyadmin - {data_para_nome_planilha}.xlsx')
nome_arquivo_csv = (f'/content/drive/MyDrive/painel de gestão/Filtro para subir no phpmyadmin - {data_para_nome_planilha}.csv')

"""Tratando a planilha"""

#caminho da planilha original
arquivo_painel = '/content/drive/MyDrive/painel de gestão/NOVO PAINEL DE GESTÃO - 01-07-24.xlsx'


#carregando a planilha com apenas dados
wb_painel = openpyxl.load_workbook(arquivo_painel, data_only=True)
ws_resumo_painel = wb_painel['RESUMO'] #Carregando a aba 'resumo'
ws_farol_painel = wb_painel['FAROL'] #Carregando a aba 'farol'


#Removendo as primeiras linhas da aba 'FAROL'
ws_farol_painel.delete_rows(1,3)


#Criando uma nova aba na planilha chamada 'resumotratado'
nova_sheet = wb_painel.create_sheet(title='resumotratado')


# Copiar e colar os valores (inteirando em cada linha e cada célula)
for row in ws_resumo_painel.iter_rows():
    for cell in row:
        nova_sheet[cell.coordinate].value = cell.value


#Removendo as primeiras linha da aba "resumotratado" que não nos serve
ws_resumotratado = wb_painel['resumotratado']
ws_resumotratado.delete_rows(1,5)


# excluindo colunas que não usaremos
ws_resumotratado.delete_cols(27)
ws_resumotratado.delete_cols(23)
ws_resumotratado.delete_cols(21)
ws_resumotratado.delete_cols(14)
ws_resumotratado.delete_cols(11, 2)
ws_resumotratado.delete_cols(8, 2)
ws_resumotratado.delete_cols(6)
ws_resumotratado.delete_cols(2, 3)


#Depois que excluirmos, vamos adicionar uma coluna nova com a data incial e final do painel
ws_resumotratado.insert_cols(2) #Inserindo coluna para data inicial
ws_resumotratado.insert_cols(3) #Inserindo coluna para data final


#Informando os valores das datas tratados ('%d/%m/%Y')
ws_resumotratado.cell(row=1, column=2).value = "inicio"
ws_resumotratado.cell(row=1, column=3).value = "fim"


#Inteirando sobre as linhas A2:A74 e B2:B74 para adicional os valores das datas
for row in range(2,75):
    ws_resumotratado.cell(row=row, column=2).value = data_inicial_formatada
    ws_resumotratado.cell(row=row, column=3).value = data_final_formatada


#Pegando as colunas de pontuação final e ranking

for row in range(1, ws_farol_painel.max_row + 1):
  copiando_valor_pontos = ws_farol_painel.cell(row=row, column=20).value
  copiando_valor_ranking = ws_farol_painel.cell(row=row, column=21).value


  #colando os calores na aba 'resumotratado'
  ws_resumotratado.cell(row=row, column=19).value = copiando_valor_pontos
  ws_resumotratado.cell(row=row, column=20).value = copiando_valor_ranking


#Removendo as primeiras linha da aba "resumotratado" que não nos serve após adicionar os valores
#da aba "farol"
ws_resumotratado.delete_rows(1)


#inteirando sobre todos os valores se ele for "ok" trocaremos para 1 e se for "nok", trocaremos por 0
for row in ws_resumotratado.iter_rows():
  for cell in row:
    if cell.value == "OK":
      cell.value = 1
    elif cell.value == "NOK":
      cell.value = 0

"""Salvando o arquivo"""

# Criar um novo workbook
nova_planilha_com_resumotratado = openpyxl.Workbook()

# Copiar a planilha 'resumotratado' para o novo workbook
nova_aba = nova_planilha_com_resumotratado.active
nova_aba.title = 'Filtro para subir no phpmyadmin'

# Copiar célula por célula
for row in ws_resumotratado.iter_rows():
    for cell in row:
        nova_aba[cell.coordinate].value = cell.value

# Salvar o novo workbook
nova_planilha_com_resumotratado.save(nome_arquivo)

df_arquivo_para_float_csv = pd.read_excel(nome_arquivo, header=None)


#colunas que vamos interar para converter em int
colunas_int = [0, 5, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19]


#colunas que vamos interar para converter em int
colunas_float = [4, 6, 13]


# Converter colunas para float
for col in colunas_float:
    df_arquivo_para_float_csv[col] = pd.to_numeric(df_arquivo_para_float_csv[col], errors='coerce')

# Converter colunas para int
for col in colunas_int:
    df_arquivo_para_float_csv[col] = pd.to_numeric(df_arquivo_para_float_csv[col], errors='coerce').fillna(0).astype(int)

# df_arquivo_para_float_csv.head(70)

# Salvar o DataFrame para um arquivo CSV sem cabeçalho e sem índices
df_arquivo_para_float_csv.to_csv(nome_arquivo_csv, index=False, header=False, float_format='%.2f')

